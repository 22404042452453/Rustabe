#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Модуль для анализа устойчивости энергосистем с использованием RastrWin.
Содержит основные функции расчета и классы для работы с генераторами.
"""

import win32com.client
import time
import csv
import json
import os
from datetime import datetime

# сообщения RastrRetCode
RastrRetCodeString = [
    "успешное завершение",
    "ошибка при проведении расчёта"]

# сообщения SyncLossCause
SyncLossCauseString = [
    "расчет завершен успешно, потери синхронизма не выявлено",
    "выявлено превышение угла по ветви значения 180",
    "выявлено превышение угла по сопротивлению генератора значения 180",
    "выявлено превышение допустимой скорости вращения одного или нескольких генераторов"]


class MultiVariable:
    """
    Класс для управления несколькими генераторами одновременно
    """
    def __init__(self, rastr, generators):
        """
        generators - список словарей с параметрами генераторов
        Каждый генератор: {
            "table": "Generator",
            "column": "P",
            "key": "Num = 51804050",
            "name": "Ген-1"
        }
        """
        self.rastr = rastr
        self.generators = []

        for gen_config in generators:
            gen_info = {
                'table': self.rastr.Tables(gen_config['table']),
                'column': None,
                'row_id': None,
                'name': gen_config.get('name', gen_config['key'])
            }

            gen_info['column'] = gen_info['table'].Cols(gen_config['column'])
            gen_info['table'].SetSel(gen_config['key'])
            gen_info['row_id'] = gen_info['table'].FindNextSel(-1)

            self.generators.append(gen_info)

    def apply(self, values):
        """
        Применяет значения мощности к генераторам
        values - список значений или одно значение (применится ко всем)
        """
        if not isinstance(values, list):
            values = [values] * len(self.generators)

        for gen_info, value in zip(self.generators, values):
            if value is not None:
                gen_info['column'].SetZ(gen_info['row_id'], value)

        self.rastr.rgm("")

    def get(self):
        """Возвращает текущие значения мощности всех генераторов"""
        return [gen['column'].Z(gen['row_id']) for gen in self.generators]

    def get_deltas(self):
        """Возвращает текущие углы всех генераторов"""
        deltas = []
        for gen in self.generators:
            delta_col = gen['table'].Cols('Delta')
            deltas.append(delta_col.Z(gen['row_id']))
        return deltas

    def get_parameters(self, param_types):
        """
        Получает значения выбранных параметров для всех генераторов

        param_types - список типов параметров ['delta', 'voltage', 'power_p', 'power_q', 'current']
        Возвращает словарь {param_type: [values]}
        """
        results = {}

        for param_type in param_types:
            if param_type == 'delta':
                results['delta'] = self.get_deltas()
            elif param_type == 'voltage':
                results['voltage'] = [gen['table'].Cols('U').Z(gen['row_id']) for gen in self.generators]
            elif param_type == 'power_p':
                results['power_p'] = [gen['table'].Cols('P').Z(gen['row_id']) for gen in self.generators]
            elif param_type == 'power_q':
                results['power_q'] = [gen['table'].Cols('Q').Z(gen['row_id']) for gen in self.generators]
            elif param_type == 'current':
                # Ток может не быть напрямую доступен, в зависимости от модели RastrWin
                try:
                    results['current'] = [gen['table'].Cols('I').Z(gen['row_id']) for gen in self.generators]
                except:
                    results['current'] = ["N/A"] * len(self.generators)

        return results


def simple_calculate(var, powers, gen_ids_to_check=None, selected_params=None):
    """
    Расчет устойчивости для заданных мощностей генераторов (простой метод)
    powers - список мощностей для каждого генератора
    gen_ids_to_check - список индексов генераторов для проверки (None = все)
    selected_params - список параметров для мониторинга
    """
    print("/---------")
    print("Расчёт устойчивости для мощностей:", powers, " МВт")

    tic_calc = time.perf_counter()
    var.apply(powers)

    ret_code = var.rastr.FWDynamic().Run()
    print(RastrRetCodeString[ret_code])

    # Проверяем все генераторы или только указанные
    if gen_ids_to_check is None:
        gen_ids_to_check = range(len(var.generators))

    SyncLossCause = 0
    comment = ""

    # Проверяем углы генераторов (основная логика устойчивости)
    for gen_idx in gen_ids_to_check:
        gen = var.generators[gen_idx]
        data = var.rastr.GetChainedGraphSnapshot("Generator", "Delta", gen['row_id'], 0)

        for (val, t) in data:
            if t < 0.1:
                base_t = t
                base_val = val
            if (val > (base_val + 180)) or (val < (base_val - 180)):
                SyncLossCause = 1
                msg = f"Генератор {gen['name']} : угол превысил 180° в момент времени {t:.3f}с"
                print(msg)
                comment = msg
                break

        if SyncLossCause == 1:
            break

    if SyncLossCause == 0:
        print("Все генераторы устойчивы")
        comment = "Все генераторы устойчивы"

    calc_time = time.perf_counter() - tic_calc
    print(f"Время расчёта {calc_time:.2f} с")
    print("\\---------")

    # Собираем дополнительные параметры если запрошены
    parameters = {}
    if selected_params:
        parameters = var.get_parameters(selected_params)

    return {
        'stable': (SyncLossCause == 0),
        'calc_time': calc_time,
        'comment': comment,
        'parameters': parameters
    }


def ems_calculate(var, powers, selected_params=None):
    """
    Расчет через EMS режим
    powers - список мощностей для каждого генератора
    selected_params - список параметров для мониторинга
    """
    print("/---------")
    print("Расчёт устойчивости для мощностей:", powers, " МВт")

    tic_calc = time.perf_counter()
    var.apply(powers)

    dyn = var.rastr.FWDynamic()
    ret_code = dyn.RunEMSmode()
    print(RastrRetCodeString[ret_code])
    print(SyncLossCauseString[dyn.SyncLossCause])
    print("промоделированное время", dyn.TimeReached)

    calc_time = time.perf_counter() - tic_calc
    print(f"Время расчёта {calc_time:.2f} с")
    print("\\---------")

    comment = dyn.ResultMessage or SyncLossCauseString[dyn.SyncLossCause]

    # Собираем дополнительные параметры если запрошены
    parameters = {}
    if selected_params:
        parameters = var.get_parameters(selected_params)

    return {
        'stable': (dyn.SyncLossCause == 0),
        'calc_time': calc_time,
        'comment': comment,
        'parameters': parameters
    }


class MultiScenarioAnalyzer:
    """
    Основной класс для проведения многовариантного анализа
    """

    def __init__(self):
        """Инициализация анализатора"""
        self.rastr = None
        self.dyn = None

    def run_analysis(self, net_file, scenario_files, scenario_folder, generators_config,
                    power_ranges, calculate_func="simple", selected_params=None,
                    progress_callback=None):
        """
        Запуск многовариантного анализа

        net_file - файл сети
        scenario_files - список файлов сценариев
        scenario_folder - папка со сценариями
        generators_config - конфигурация генераторов
        power_ranges - диапазоны мощностей [(min, max), ...]
        calculate_func - метод расчета ("simple" или "ems")
        selected_params - параметры для мониторинга
        progress_callback - функция для обновления прогресса
        """
        # Подключаемся к RastrWin если еще не подключены
        if self.rastr is None:
            try:
                self.rastr = win32com.client.Dispatch("Astra.Rastr")
                self.dyn = self.rastr.FWDynamic()
                print("Подключен модуль RastrWin")
            except Exception as e:
                raise Exception(f"Не удалось подключиться к RastrWin: {e}")

        results = {}
        total_scenarios = len(scenario_files)
        completed = 0

        for scenario_file in scenario_files:
            if progress_callback:
                progress_callback(f"Обработка сценария: {scenario_file}", (completed / total_scenarios) * 100)

            print(f"\n{'='*50}")
            print(f"СЦЕНАРИЙ: {scenario_file}")
            print(f"{'='*50}")

            # Загрузка файлов режима
            scenario_path = os.path.join(scenario_folder, scenario_file)
            print("Загрузка файлов режима")

            try:
                self.rastr.Load(1, net_file, "Шаблон/динамика.rst")
                self.rastr.Load(1, scenario_path, "Шаблон/сценарий.scn")
                self.rastr.Load(1, "", "Шаблон/автоматика.dfw")
            except Exception as e:
                print(f"Ошибка загрузки файлов для сценария {scenario_file}: {e}")
                continue

            # Создание объекта для управления генераторами
            var = MultiVariable(self.rastr, generators_config)

            # Настройка снапшотов
            table = self.rastr.Tables("com_dynamics")
            column = table.Cols("SnapMaxCount")
            column.SetZ(0, 1)

            # Тестирование различных комбинаций мощностей
            print("\nТестирование устойчивости...")

            scenario_results = []

            # Простой перебор граничных значений
            test_cases = [
                [power_ranges[i][0] for i in range(len(power_ranges))],  # Минимумы
                [power_ranges[i][1] for i in range(len(power_ranges))],  # Максимумы
                [(power_ranges[i][0] + power_ranges[i][1])/2 for i in range(len(power_ranges))]  # Средние
            ]

            for powers in test_cases:
                if calculate_func == "ems":
                    result = ems_calculate(var, powers, selected_params)
                else:
                    result = simple_calculate(var, powers, selected_params=selected_params)

                scenario_results.append({
                    'powers': powers,
                    'stable': result['stable'],
                    'calc_time': result['calc_time'],
                    'comment': result['comment'],
                    'parameters': result.get('parameters', {})
                })

            results[scenario_file] = scenario_results
            completed += 1

            if progress_callback:
                progress_callback(f"Завершен сценарий: {scenario_file}", (completed / total_scenarios) * 100)

        if progress_callback:
            progress_callback("Анализ завершен", 100)

        return results

    def export_to_csv(self, results, generators_config, filename="results.csv"):
        """Экспорт результатов в CSV"""
        print(f"\nЭкспорт в CSV: {filename}")

        with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile, delimiter=';')

            # Заголовок
            header = ["Сценарий"]
            for i, gen in enumerate(generators_config):
                header.append(f"Ген {i+1} (МВт)")
            header.extend(["Результат", "Время расчета (с)", "Комментарий"])

            writer.writerow(header)

            # Данные
            for scenario, scenario_results in results.items():
                for result in scenario_results:
                    row = [scenario]
                    for power in result['powers']:
                        row.append(str(power))
                    status = "Устойчиво" if result['stable'] else "Неустойчиво"
                    row.append(status)
                    row.append(f"{result.get('calc_time', 0):.2f}")
                    row.append(result.get('comment', ''))
                    writer.writerow(row)

        print("CSV файл сохранен")

    def export_to_excel(self, results, generators_config, filename="results.xlsx"):
        """Экспорт результатов в Excel с несколькими вкладками"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            print(f"\nЭкспорт в Excel: {filename}")

            wb = openpyxl.Workbook()

            # Стили
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            stable_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            unstable_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center")

            # === Вкладка "Мощность" ===
            ws_power = wb.active
            ws_power.title = "Мощность"

            headers_power = ["Сценарий"]
            for i, gen in enumerate(generators_config):
                headers_power.append(f"Ген {i+1} ({gen.get('name', '')})")
            headers_power.extend(["Результат", "Время расчета (с)", "Комментарий"])

            for col, header in enumerate(headers_power, 1):
                cell = ws_power.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            row_num = 2
            for scenario, scenario_results in results.items():
                for result in scenario_results:
                    ws_power.cell(row=row_num, column=1, value=scenario)

                    for col, power in enumerate(result['powers'], 2):
                        ws_power.cell(row=row_num, column=col, value=power)

                    status = "Устойчиво" if result['stable'] else "Неустойчиво"
                    status_cell = ws_power.cell(row=row_num, column=len(result['powers']) + 2, value=status)
                    status_cell.alignment = center_align

                    if result['stable']:
                        status_cell.fill = stable_fill
                    else:
                        status_cell.fill = unstable_fill

                    ws_power.cell(row=row_num, column=len(result['powers']) + 3,
                                 value=round(result.get('calc_time', 0), 2))

                    ws_power.cell(row=row_num, column=len(result['powers']) + 4,
                                 value=result.get('comment', ''))

                    row_num += 1

            # === Вкладка "Угол ротора" ===
            ws_delta = wb.create_sheet("Угол ротора")
            headers_delta = ["Сценарий"]
            for i, gen in enumerate(generators_config):
                headers_delta.append(f"Ген {i+1} ({gen.get('name', '')})")
            headers_delta.append("Результат")

            for col, header in enumerate(headers_delta, 1):
                cell = ws_delta.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            row_num = 2
            for scenario, scenario_results in results.items():
                for result in scenario_results:
                    ws_delta.cell(row=row_num, column=1, value=scenario)

                    if 'parameters' in result and 'delta' in result['parameters']:
                        deltas = result['parameters']['delta']
                        for col, delta in enumerate(deltas, 2):
                            ws_delta.cell(row=row_num, column=col, value=round(delta, 2))
                    else:
                        for col in range(2, len(generators_config) + 2):
                            ws_delta.cell(row=row_num, column=col, value="Н/Д")

                    status = "Устойчиво" if result['stable'] else "Неустойчиво"
                    status_cell = ws_delta.cell(row=row_num, column=len(generators_config) + 2, value=status)
                    status_cell.alignment = center_align

                    if result['stable']:
                        status_cell.fill = stable_fill
                    else:
                        status_cell.fill = unstable_fill

                    row_num += 1

            # === Вкладка "Напряжение" ===
            ws_voltage = wb.create_sheet("Напряжение")
            headers_voltage = ["Сценарий"]
            for i, gen in enumerate(generators_config):
                headers_voltage.append(f"Ген {i+1} ({gen.get('name', '')})")
            headers_voltage.append("Результат")

            for col, header in enumerate(headers_voltage, 1):
                cell = ws_voltage.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            row_num = 2
            for scenario, scenario_results in results.items():
                for result in scenario_results:
                    ws_voltage.cell(row=row_num, column=1, value=scenario)

                    if 'parameters' in result and 'voltage' in result['parameters']:
                        voltages = result['parameters']['voltage']
                        for col, voltage in enumerate(voltages, 2):
                            ws_voltage.cell(row=row_num, column=col, value=round(voltage, 2))
                    else:
                        for col in range(2, len(generators_config) + 2):
                            ws_voltage.cell(row=row_num, column=col, value="Н/Д")

                    status = "Устойчиво" if result['stable'] else "Неустойчиво"
                    status_cell = ws_voltage.cell(row=row_num, column=len(generators_config) + 2, value=status)
                    status_cell.alignment = center_align

                    if result['stable']:
                        status_cell.fill = stable_fill
                    else:
                        status_cell.fill = unstable_fill

                    row_num += 1

            # === Вкладка "Ток" ===
            ws_current = wb.create_sheet("Ток")
            headers_current = ["Сценарий"]
            for i, gen in enumerate(generators_config):
                headers_current.append(f"Ген {i+1} ({gen.get('name', '')})")
            headers_current.append("Результат")

            for col, header in enumerate(headers_current, 1):
                cell = ws_current.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            row_num = 2
            for scenario, scenario_results in results.items():
                for result in scenario_results:
                    ws_current.cell(row=row_num, column=1, value=scenario)

                    if 'parameters' in result and 'current' in result['parameters']:
                        currents = result['parameters']['current']
                        for col, current in enumerate(currents, 2):
                            if isinstance(current, str):
                                ws_current.cell(row=row_num, column=col, value=current)
                            else:
                                ws_current.cell(row=row_num, column=col, value=round(current, 2))
                    else:
                        for col in range(2, len(generators_config) + 2):
                            ws_current.cell(row=row_num, column=col, value="Н/Д")

                    status = "Устойчиво" if result['stable'] else "Неустойчиво"
                    status_cell = ws_current.cell(row=row_num, column=len(generators_config) + 2, value=status)
                    status_cell.alignment = center_align

                    if result['stable']:
                        status_cell.fill = stable_fill
                    else:
                        status_cell.fill = unstable_fill

                    row_num += 1

            # Автоподбор ширины колонок для всех листов
            for ws in [ws_power, ws_delta, ws_voltage, ws_current]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Сохранение
            wb.save(filename)
            print("Excel файл сохранен с вкладками: Мощность, Угол ротора, Напряжение, Ток")

        except ImportError:
            print("Excel экспорт недоступен. Установите openpyxl: pip install openpyxl")

    def export_to_json(self, results, generators_config, filename="results.json"):
        """Экспорт результатов в JSON"""
        print(f"\nЭкспорт в JSON: {filename}")

        export_data = {
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'generators': generators_config,
            'results': {}
        }

        for scenario, scenario_results in results.items():
            export_data['results'][scenario] = [
                {
                    'powers': result['powers'],
                    'stable': result['stable'],
                    'calc_time': result.get('calc_time', 0),
                    'comment': result.get('comment', ''),
                    'parameters': result.get('parameters', {})
                }
                for result in scenario_results
            ]

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)

        print("JSON файл сохранен")
